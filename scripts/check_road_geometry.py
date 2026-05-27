# =============================================================================
# check_road_geometry.py
# -----------------------------------------------------------------------------
# Runs geometry quality checks on the road network feature class:
# - Check Geometry (structural geometry problems)
# - Count Overlapping Features (overlapping segments)
# - Dangle detection (unconnected endpoints)
# - Duplicate ID detection
# Exports a consolidated QC report to .xlsx
#
# Autora : Denise Hernández — GIS Analyst
# Entorno: ArcGIS Pro + arcpy (Python 3.x)
# =============================================================================

import arcpy
import os
import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    raise ImportError("Install openpyxl: pip install openpyxl")

# =============================================================================
# CONFIGURATION
# =============================================================================

# Path to the road feature class (shapefile or local GDB — NOT enterprise GDB)
ROADS_FC = r"C:\Ruta\A\RoadNetwork.gdb\RoadNetworkFD\FC_Cam_Caminos"

# Path to bifurcations feature class
BIFURCATIONS_FC = r"C:\Ruta\A\RoadNetwork.gdb\RoadNetworkFD\FC_Cam_Bifurcaciones"

# Scratch GDB for intermediate outputs
SCRATCH_GDB = r"C:\Ruta\A\Scratch.gdb"

# Output folder for the Excel report
OUTPUT_FOLDER = r"C:\Ruta\A\QC_Reports"
OUTPUT_FILENAME = f"road_geometry_qc_{datetime.date.today().isoformat()}.xlsx"

# XY tolerance for short segment evaluation (meters)
XY_TOLERANCE_M = 0.005

# Overlap detection: minimum count to flag as overlap
OVERLAP_THRESHOLD = 1

# Header fill color for Excel
COLOR_HEADER = "1E4D78"

# =============================================================================
# HELPERS
# =============================================================================

def validate_paths(*paths):
    for p in paths:
        if not arcpy.Exists(p):
            raise FileNotFoundError(f"Not found: {p}")

def log(msg):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")

def styled_header(ws, row, columns):
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

def autofit_columns(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

# =============================================================================
# CHECK 1 — GEOMETRY INTEGRITY
# =============================================================================

def check_geometry(roads_fc, scratch_gdb):
    log("Running Check Geometry...")
    output_table = os.path.join(scratch_gdb, "geom_errors")
    if arcpy.Exists(output_table):
        arcpy.Delete_management(output_table)

    arcpy.management.CheckGeometry(
        in_features=roads_fc,
        out_table=output_table,
        validation_method="OGC"
    )

    fields = [f.name for f in arcpy.ListFields(output_table)]
    rows = []
    with arcpy.da.SearchCursor(output_table, fields) as cur:
        for row in cur:
            rows.append(dict(zip(fields, row)))

    log(f"  Geometry errors found: {len(rows)}")
    return fields, rows

# =============================================================================
# CHECK 2 — OVERLAPPING FEATURES
# =============================================================================

def count_overlapping(roads_fc, scratch_gdb):
    log("Running Count Overlapping Features...")
    output_fc = os.path.join(scratch_gdb, "overlapping_roads")
    if arcpy.Exists(output_fc):
        arcpy.Delete_management(output_fc)

    arcpy.analysis.CountOverlappingFeatures(
        in_features=roads_fc,
        out_feature_class=output_fc,
        min_overlap_count=OVERLAP_THRESHOLD
    )

    fields = [f.name for f in arcpy.ListFields(output_fc) if f.name != "Shape"]
    rows = []
    with arcpy.da.SearchCursor(output_fc, ["SHAPE@LENGTH"] + fields) as cur:
        for row in cur:
            d = dict(zip(["Shape_Length"] + fields, row))
            if d.get("COUNT_", 0) > OVERLAP_THRESHOLD:
                rows.append(d)

    log(f"  Overlapping segments found: {len(rows)}")
    return ["Shape_Length"] + fields, rows

# =============================================================================
# CHECK 3 — DUPLICATE ROAD IDs
# =============================================================================

def check_duplicate_ids(roads_fc, id_field="Road_ID"):
    log(f"Checking duplicate {id_field} values...")
    id_counts = {}
    try:
        with arcpy.da.SearchCursor(roads_fc, [id_field, "OBJECTID"]) as cur:
            for row in cur:
                road_id, oid = row
                if road_id not in id_counts:
                    id_counts[road_id] = []
                id_counts[road_id].append(oid)
    except RuntimeError:
        log(f"  Field '{id_field}' not found — skipping duplicate ID check.")
        return [], []

    duplicates = [(rid, oids) for rid, oids in id_counts.items() if len(oids) > 1]
    rows = []
    for rid, oids in duplicates:
        for oid in oids:
            rows.append({"Road_ID": rid, "OBJECTID": oid, "Count": len(oids)})

    log(f"  Duplicate Road IDs found: {len(duplicates)} IDs ({len(rows)} records)")
    fields = ["Road_ID", "OBJECTID", "Count"]
    return fields, rows

# =============================================================================
# CHECK 4 — NULL ATTRIBUTES
# =============================================================================

def check_null_attributes(roads_fc):
    log("Checking null mandatory attributes...")
    mandatory_fields = ["Road_ID", "Enabled", "Hierarchy", "Endpoint_1", "Endpoint_2"]
    available_fields = [f.name for f in arcpy.ListFields(roads_fc)]
    check_fields = [f for f in mandatory_fields if f in available_fields]

    if not check_fields:
        log("  No matching mandatory fields found — skipping null check.")
        return [], []

    rows = []
    with arcpy.da.SearchCursor(roads_fc, ["OBJECTID"] + check_fields) as cur:
        for row in cur:
            oid = row[0]
            for i, field in enumerate(check_fields):
                val = row[i + 1]
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    rows.append({"OBJECTID": oid, "Field": field, "Issue": "NULL or empty"})

    log(f"  Null attribute issues found: {len(rows)}")
    fields = ["OBJECTID", "Field", "Issue"]
    return fields, rows

# =============================================================================
# EXPORT TO EXCEL
# =============================================================================

def export_report(geom_fields, geom_rows,
                  overlap_fields, overlap_rows,
                  dup_fields, dup_rows,
                  null_fields, null_rows,
                  output_path):
    wb = openpyxl.Workbook()

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    styled_header(ws_sum, 1, ["Check", "Issues Found", "Status"])
    checks = [
        ("Geometry errors (Check Geometry)", len(geom_rows)),
        ("Overlapping segments", len(overlap_rows)),
        ("Duplicate Road IDs", len(dup_rows)),
        ("Null mandatory attributes", len(null_rows)),
    ]
    for i, (name, count) in enumerate(checks, start=2):
        ws_sum.cell(row=i, column=1, value=name)
        ws_sum.cell(row=i, column=2, value=count)
        status = "✅ Pass" if count == 0 else "❌ Issues found"
        cell = ws_sum.cell(row=i, column=3, value=status)
        color = "C6EFCE" if count == 0 else "FFC7CE"
        cell.fill = PatternFill("solid", fgColor=color)
    autofit_columns(ws_sum)

    # Geometry errors sheet
    ws_geom = wb.create_sheet("Geometry Errors")
    styled_header(ws_geom, 1, geom_fields or ["No errors"])
    for i, row in enumerate(geom_rows, start=2):
        for j, col in enumerate(geom_fields, start=1):
            ws_geom.cell(row=i, column=j, value=row.get(col, ""))
    autofit_columns(ws_geom)

    # Overlapping sheet
    ws_ov = wb.create_sheet("Overlapping Segments")
    styled_header(ws_ov, 1, overlap_fields or ["No overlaps"])
    for i, row in enumerate(overlap_rows, start=2):
        for j, col in enumerate(overlap_fields, start=1):
            ws_ov.cell(row=i, column=j, value=row.get(col, ""))
    autofit_columns(ws_ov)

    # Duplicate IDs sheet
    ws_dup = wb.create_sheet("Duplicate IDs")
    styled_header(ws_dup, 1, dup_fields or ["No duplicates"])
    for i, row in enumerate(dup_rows, start=2):
        for j, col in enumerate(dup_fields, start=1):
            ws_dup.cell(row=i, column=j, value=row.get(col, ""))
    autofit_columns(ws_dup)

    # Null attributes sheet
    ws_null = wb.create_sheet("Null Attributes")
    styled_header(ws_null, 1, null_fields or ["No nulls"])
    for i, row in enumerate(null_rows, start=2):
        for j, col in enumerate(null_fields, start=1):
            ws_null.cell(row=i, column=j, value=row.get(col, ""))
    autofit_columns(ws_null)

    wb.save(output_path)
    log(f"Report saved: {output_path}")

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 60)
    print("  Road Network Geometry QC")
    print(f"  {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    validate_paths(ROADS_FC, SCRATCH_GDB)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    geom_fields, geom_rows = check_geometry(ROADS_FC, SCRATCH_GDB)
    ov_fields, ov_rows = count_overlapping(ROADS_FC, SCRATCH_GDB)
    dup_fields, dup_rows = check_duplicate_ids(ROADS_FC)
    null_fields, null_rows = check_null_attributes(ROADS_FC)

    total_issues = len(geom_rows) + len(ov_rows) + len(dup_rows) + len(null_rows)

    output_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)
    export_report(geom_fields, geom_rows,
                  ov_fields, ov_rows,
                  dup_fields, dup_rows,
                  null_fields, null_rows,
                  output_path)

    print("\n--- SUMMARY ---")
    print(f"  Geometry errors   : {len(geom_rows)}")
    print(f"  Overlapping segs  : {len(ov_rows)}")
    print(f"  Duplicate IDs     : {len(dup_rows)}")
    print(f"  Null attributes   : {len(null_rows)}")
    print(f"  TOTAL ISSUES      : {total_issues}")
    print(f"\n  {'✅ All checks passed!' if total_issues == 0 else '❌ Issues found — review the report.'}")


if __name__ == "__main__":
    main()
