# =============================================================================
# blockage_analysis.py
# -----------------------------------------------------------------------------
# Detects road barriers (gates, cattle guards, blockages) with Enabled = NO
# that are affecting routing in the road network.
# Performs spatial join between barriers and roads, detects duplicate geometry,
# and exports a blockage report to .xlsx.
#
# Autora : Denise Hernández — GIS Analyst
# Entorno: ArcGIS Pro + arcpy (Python 3.x)
# =============================================================================

import arcpy
import os
import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    raise ImportError("Install openpyxl: pip install openpyxl")

# =============================================================================
# CONFIGURATION
# =============================================================================

ROADS_FC      = r"C:\Ruta\A\RoadNetwork.gdb\RoadNetworkFD\FC_Cam_Caminos"
BLOCKAGES_FC  = r"C:\Ruta\A\RoadNetwork.gdb\RoadNetworkFD\FC_Cam_Bloqueos"
SCRATCH_GDB   = r"C:\Ruta\A\Scratch.gdb"
OUTPUT_FOLDER = r"C:\Ruta\A\Reports"
OUTPUT_FILE   = f"blockage_report_{datetime.date.today().isoformat()}.xlsx"

# Spatial join search radius (meters)
SPATIAL_JOIN_RADIUS_M = "20 Meters"

# Field names (adjust to match your schema)
FIELD_ENABLED      = "Enabled"
FIELD_ROAD_ID      = "Road_ID"
FIELD_BLOCKAGE_ID  = "Blockage_ID"
FIELD_TYPE         = "Type"

COLOR_HEADER    = "1E4D78"
COLOR_BLOCKED   = "FFC7CE"   # red — disabled barriers
COLOR_CLEAR     = "C6EFCE"   # green — enabled barriers
COLOR_DUPLICATE = "FFEB9C"   # yellow — duplicate geometry

# =============================================================================
# HELPERS
# =============================================================================

def log(msg):
    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}")

def header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center")

def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

# =============================================================================
# STEP 1 — SPATIAL JOIN: BARRIERS × ROADS
# =============================================================================

def spatial_join_barriers_roads(blockages_fc, roads_fc, scratch_gdb):
    log("Spatial join: barriers × roads...")
    output = os.path.join(scratch_gdb, "barriers_joined")
    if arcpy.Exists(output):
        arcpy.Delete_management(output)

    arcpy.analysis.SpatialJoin(
        target_features=blockages_fc,
        join_features=roads_fc,
        out_feature_class=output,
        join_operation="JOIN_ONE_TO_ONE",
        join_type="KEEP_ALL",
        match_option="CLOSEST",
        search_radius=SPATIAL_JOIN_RADIUS_M
    )

    fields = [f.name for f in arcpy.ListFields(output)
              if f.name not in ("Shape", "Shape_Length", "Shape_Area")]

    rows = []
    with arcpy.da.SearchCursor(output, fields) as cur:
        for row in cur:
            rows.append(dict(zip(fields, row)))

    log(f"  Joined records: {len(rows)}")
    return fields, rows

# =============================================================================
# STEP 2 — FILTER DISABLED BARRIERS
# =============================================================================

def filter_disabled(rows, enabled_field=FIELD_ENABLED):
    disabled = [r for r in rows if str(r.get(enabled_field, "")).upper() == "NO"]
    log(f"  Disabled barriers (Enabled=NO): {len(disabled)}")
    return disabled

# =============================================================================
# STEP 3 — DETECT DUPLICATE GEOMETRY
# =============================================================================

def detect_duplicate_geometry(blockages_fc, scratch_gdb):
    log("Detecting duplicate barrier geometry...")
    output = os.path.join(scratch_gdb, "duplicate_barriers")
    if arcpy.Exists(output):
        arcpy.Delete_management(output)

    arcpy.management.FindIdentical(
        in_dataset=blockages_fc,
        out_dataset=output,
        fields="Shape",
        output_record_option="ONLY_DUPLICATES"
    )

    rows = []
    try:
        with arcpy.da.SearchCursor(output, ["IN_FID", "FEAT_SEQ"]) as cur:
            for row in cur:
                rows.append({"IN_FID": row[0], "FEAT_SEQ": row[1]})
    except Exception:
        pass

    log(f"  Duplicate barrier geometries found: {len(rows)}")
    return rows

# =============================================================================
# STEP 4 — EXPORT REPORT
# =============================================================================

def export_report(all_rows, disabled_rows, dup_rows, fields, output_path):
    wb = openpyxl.Workbook()

    # Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    for cell in [ws_sum.cell(1, 1, "Metric"),
                 ws_sum.cell(1, 2, "Count"),
                 ws_sum.cell(1, 3, "Status")]:
        header_style(cell)

    summary = [
        ("Total barriers analyzed", len(all_rows), None),
        ("Disabled barriers (Enabled=NO)", len(disabled_rows), "FFC7CE" if disabled_rows else "C6EFCE"),
        ("Duplicate geometries detected", len(dup_rows), "FFEB9C" if dup_rows else "C6EFCE"),
    ]
    for i, (metric, count, color) in enumerate(summary, start=2):
        ws_sum.cell(row=i, column=1, value=metric)
        ws_sum.cell(row=i, column=2, value=count)
        status = "⚠️ Requires action" if count > 0 and color else "✅ OK"
        cell = ws_sum.cell(row=i, column=3, value=status)
        if color:
            cell.fill = PatternFill("solid", fgColor=color)
    autofit(ws_sum)

    # All barriers
    ws_all = wb.create_sheet("All Barriers")
    for col_idx, col_name in enumerate(fields, start=1):
        header_style(ws_all.cell(1, col_idx, col_name))
    for i, row in enumerate(all_rows, start=2):
        enabled = str(row.get(FIELD_ENABLED, "")).upper()
        fill = PatternFill("solid", fgColor=COLOR_BLOCKED if enabled == "NO" else COLOR_CLEAR)
        for j, col in enumerate(fields, start=1):
            cell = ws_all.cell(row=i, column=j, value=row.get(col, ""))
            cell.fill = fill
    autofit(ws_all)

    # Disabled barriers
    ws_dis = wb.create_sheet("Disabled Barriers")
    for col_idx, col_name in enumerate(fields, start=1):
        header_style(ws_dis.cell(1, col_idx, col_name))
    for i, row in enumerate(disabled_rows, start=2):
        for j, col in enumerate(fields, start=1):
            ws_dis.cell(row=i, column=j, value=row.get(col, ""))
    autofit(ws_dis)

    # Duplicate geometries
    ws_dup = wb.create_sheet("Duplicate Geometry")
    for col_idx, col_name in enumerate(["IN_FID", "FEAT_SEQ"], start=1):
        header_style(ws_dup.cell(1, col_idx, col_name))
    for i, row in enumerate(dup_rows, start=2):
        ws_dup.cell(row=i, column=1, value=row["IN_FID"]).fill = PatternFill("solid", fgColor=COLOR_DUPLICATE)
        ws_dup.cell(row=i, column=2, value=row["FEAT_SEQ"])
    autofit(ws_dup)

    wb.save(output_path)
    log(f"Report saved: {output_path}")

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 60)
    print("  Road Blockage Analysis")
    print(f"  {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    for path in [ROADS_FC, BLOCKAGES_FC, SCRATCH_GDB]:
        if not arcpy.Exists(path):
            raise FileNotFoundError(f"Not found: {path}")

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    fields, all_rows  = spatial_join_barriers_roads(BLOCKAGES_FC, ROADS_FC, SCRATCH_GDB)
    disabled_rows     = filter_disabled(all_rows)
    dup_rows          = detect_duplicate_geometry(BLOCKAGES_FC, SCRATCH_GDB)

    output_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILE)
    export_report(all_rows, disabled_rows, dup_rows, fields, output_path)

    print("\n--- RESULTS ---")
    print(f"  Total barriers    : {len(all_rows)}")
    print(f"  Disabled (NO)     : {len(disabled_rows)}")
    print(f"  Duplicate geom.   : {len(dup_rows)}")

    if disabled_rows:
        print("\n  ⚠️  Disabled barriers found — coordinate with security team.")
        print("  Update Enabled field after field confirmation.")
    else:
        print("\n  ✅ No disabled barriers detected.")


if __name__ == "__main__":
    main()
